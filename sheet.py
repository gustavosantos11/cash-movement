import openpyxl as xl
from pathlib import Path


def create(doc_value, hist_value, ent_value, sai_value):
    def add_line(sheet, sales):
        try:
            if not doc_value and not hist_value and not ent_value and not sai_value:
                return False

            n = int(doc_value)
            his = hist_value
            ent = float(ent_value) if ent_value else 0.0
            exi = float(sai_value) if sai_value else 0.0

            past_sale = sales[-1] if sales else 0.0
            bal = balance(ent, exi, past_sale, n)

            sales.append(bal)
            sheet.append([n, his, ent, exi, bal])
            print("✅ Dados adicionados com sucesso!")
            return True

        except ValueError:
            raise ValueError("❌ entrada e saída devem ser números válidos.")
        except Exception as e:
            raise e

    def balance(entraces, exits, past_sale, document_number):
        if document_number == 1:
            return entraces - exits
        elif entraces == 0:
            return past_sale - exits
        elif exits == 0:
            return past_sale + entraces
        else:
            return past_sale + entraces - exits

    archive_name = Path('Movimento_de_caixa.xlsx')

    if archive_name.exists():
        book = xl.load_workbook(archive_name)
        sheet = book.active
        sales = [
            row[4] for row in sheet.iter_rows(
                min_row=2, values_only=True
            ) if row[4] is not None
        ]
    else:
        book = xl.Workbook()
        sheet = book.active
        sheet.title = "Movimento de caixa"
        sheet.append(["Doc. N°", "Histórico", "entracess", "Saídas", "Saldo"])
        sales = []

    added_lines = add_line(sheet, sales)

    if added_lines:
        book.save(archive_name)
